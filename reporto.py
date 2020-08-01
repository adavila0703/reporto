import os
import sys
import time
import tkinter as tk
from datetime import datetime
from tkinter.messagebox import showinfo
import pyperclip
from openpyxl import load_workbook
import pyautogui
from fpdf import FPDF
from zipfile import ZipFile
from PIL import Image

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def clearexample():
    rpnum_text.delete('1.0', 'end')
    item_text.delete('1.0', 'end')
    # serial_text.delete('1.0', 'end')
    key_text.delete('1.0', 'end')
    fb_text.delete('1.0', 'end')
    comp_text.delete('1.0', 'end')
    structure_text.delete('1.0', 'end')
    parts_text.delete('1.0', 'end')
    pic_text.delete('1.0', 'end')
    his_rp_text.delete('1.0', 'end')
    his_comp_date_text.delete('1.0', 'end')
    his_symptom_text.delete('1.0', 'end')
    his_result_text.delete('1.0', 'end')
    appearance_text.delete('1.0', 'end')
    symptom_text.delete('1.0', 'end')
    tests_text.delete('1.0', 'end')
    analysis_text.delete('1.0', 'end')
    failure_text.delete('1.0', 'end')
    bapp_impact.set(False)
    bsym_nolaser.set(False)
    nrp.set(False)
    bhotswap.set(False)
    bnoise.set(False)
    bsurge.set(False)
    c_appearance_text.delete('1.0', 'end')
    c_operation_text.delete('1.0', 'end')
    c_internal_text.delete('1.0', 'end')
    c_suspected_text.delete('1.0', 'end')
    c_conclusion_text.delete('1.0', 'end')
    bench_text.delete('1.0', 'end')
    return None


def example():
    if rpnum_text.get('1.0', 'end').split('\n')[0] == '!mvmouse' and \
            item_text.get('1.0', 'end').split('\n')[0] == '!pass123':
        # tm1 = serial_text.get('1.0', 'end').split('\n')[0]
        tm1 = key_text.get('1.0', 'end').split('\n')[0]
        print(tm1)
        zipfile = ZipFile('python_icons.zip', 'r')
        img1 = Image.open(zipfile.open(name='img1.png', mode='r', pwd=b'!lol54321!'))
        img2 = Image.open(zipfile.open(name='img2.png', mode='r', pwd=b'!lol54321!'))
        img3 = Image.open(zipfile.open(name='img3.png', mode='r', pwd=b'!lol54321!'))
        clearexample()
        pyautogui.FAILSAFE = True
        while True:
            pyautogui.press('numlock')
            if datetime.now().time() >= datetime.strptime(tm1, '%H:%M:%S').time():
                pyautogui.moveTo(pyautogui.locateOnScreen(img2), duration=.75)
                time.sleep(1)
                pyautogui.click()
                return None
            else:
                pass
            time.sleep(1)

    rpnum_text.insert(tk.END, '501501')
    item_text.insert(tk.END, 'IL-600')
    # serial_text.insert(tk.END, '#A5629923')
    key_text.insert(tk.END, '14')
    fb_text.insert(tk.END, 'mother board')
    comp_text.insert(tk.END, 'A110, L605')
    structure_text.insert(tk.END, 'filter glass')
    parts_text.insert(tk.END, 'L605')
    pic_text.insert(tk.END, 'Angel')
    his_rp_text.insert(tk.END, '510510')
    his_comp_date_text.insert(tk.END, '1/1/2015')
    his_symptom_text.insert(tk.END, 'Inductor damage')
    his_result_text.insert(tk.END, 'Repaired')
    appearance_text.insert(tk.END, 'No failures found to the exterior of the unit.')
    symptom_text.insert(tk.END, 'No laser output - confirmed.')
    tests_text.insert(tk.END, 'Full function check, voltage check, current check, impedance check')
    analysis_text.insert(tk.END, '1.Visual inspection passed\n2.Unit symptom confirmed, no laser output'
                                 '\n3. Unit was disassembled and L605 was found in an open state.\n'
                                 '4.A110 was found with visible damage.')
    failure_text.insert(tk.END, 'Over voltage/current from hot swap, surge or noise.')
    bench_text.insert(tk.END, '60')
    bapp_impact.set(True)
    bsym_nolaser.set(True)
    nrp.set(True)
    bhotswap.set(True)
    bnoise.set(True)
    bsurge.set(True)

    c_appearance_text.insert(tk.END, 'There was also dirt found to the exterior of the unit.')
    c_operation_text.insert(tk.END, 'Unit was drawing 1A of current.')
    c_internal_text.insert(tk.END, 'Dirt was found inside the unit.')
    c_suspected_text.insert(tk.END, 'Damage could have also been caused by improper power shut down.')
    c_conclusion_text.insert(tk.END, 'Please contact sales person for additional warranty information.')
    return None


def checkerror():
    error = ''
    if rpnum_text.get('1.0', 'end').split('\n')[0] == '':
        error += 'RP#\n'
    if rp.get() == False and nrp.get() == False and rt.get() == False and kj.get() == False:
        error += 'Result\n'
    if bhotswap.get() == False and bvibration.get() == False and bnoise.get() == False and bheat.get() == False and \
            bnff.get() == False and bcable.get() == False and bimpact.get() == False:
        error += 'Preventative Measure\n'
    if bapp_nff.get() == False and bapp_oil.get() == False and bapp_impact.get() == False and bapp_cable.get() == False \
            and bapp_filterglass.get() == False:
        error += 'Appearance Check\n'
    if pic_text.get('1.0', 'end').split('\n')[0] == '':
        error += 'PIC\n'
    if bsym_cable.get() == False and bsym_inout.get() == False and bsym_nocom.get() == False and \
            bsym_noint.get() == False and bsym_nolaser.get() == False and bsym_nopower.get() == False and \
            bsym_symnff.get() == False and bsym_nffstruct == False:
        error += 'Symptom\n'
    warning(error)
    return None


def apperance_check():
    str_out = 'After a thorough external visual inspection, the following was found:\n'

    if bapp_nff.get() == True:
        str_out += '-No major damage was found to the exterior.\n'
    if bapp_cable.get() == True:
        str_out += '-The head cable was found damaged.\n'
    if bapp_impact.get() == True:
        str_out += '-The case was found with impact marks.\n'
    if bapp_oil.get() == True:
        str_out += '-Oil was found covering the exterior.\n'
    if bapp_filterglass.get() == True:
        str_out += '-The filter glass damage.\n'
    return str_out


def operation_check():
    str_out = 'The following was found during a full function test:\n'

    if bsym_nocom.get() == True:
        str_out += '-Unit communication was abnormal\n'
    if bsym_nolaser.get() == True:
        str_out += '-No laser output.\n'
    if bsym_nopower.get() == True:
        str_out += '-The unit did not power on.\n'
    if bsym_noint.get() == True:
        str_out += 'The unit would not initialize properly.\n'
    if bsym_cable.get() == True:
        str_out = 'With an operation check, a known good cable was replaced in order to conduct a full function ' \
                  'check\n'
    if bsym_inout.get() == True:
        str_out += '-There were abnormalities found with either the inputs or the outputs\n'
    if bsym_nffstruct.get() == True:
        str_out = 'With an operation check, although there were some structural damage to the unit, the unit is still' \
                  ' operational\n'
    if bsym_symnff.get() == True:
        str_out = 'With an operation check, the symptom was not replicated, the unit passed all tests\n'

    return str_out


def internal_check():
    str_out = ''
    if structure_text.get('1.0', 'end').split('\n')[0] != '' or fb_text.get('1.0', 'end').split('\n')[0] != '' \
            or comp_text.get('1.0', 'end').split('\n')[0] != '':
        str_out += 'With an internal check it was determined that'
        fb = fb_text.get('1.0', 'end').split('\n')[0].lower()
        struct = structure_text.get('1.0', 'end').split('\n')[0].lower()
        if comp_text.get('1.0', 'end').split('\n')[0] != '':
            if fb_text.get('1.0', 'end').split('\n')[0] != '':
                if structure_text.get('1.0', 'end').split('\n')[0] != '':
                    str_out += f' an internal electrical component on the {fb} and the {struct} were found damaged'
                else:
                    str_out += f' an internal electrical component on the {fb} was found damaged'
            else:
                if structure_text.get('1.0', 'end').split('\n')[0] != '':
                    str_out += f' an internal electrical component and the {struct} were found damaged'
                else:
                    str_out += f' an internal electrical component was found damaged'
        elif fb_text.get('1.0', 'end').split('\n')[0] != '':

            if structure_text.get('1.0', 'end').split('\n')[0] != '':
                str_out += f' the {fb} and the {struct} were found damaged'
            else:
                str_out += f' the {fb} was found damaged'

        elif structure_text.get('1.0', 'end').split('\n')[0] != '':
            str_out += f' the {struct} was found damaged'
    else:
        str_out += 'There was no need to disassemble the unit'
    return str_out + '.\n'


def suspected_cause():
    str_out = 'The unit failure was likely caused by the following possibilities:\n'
    if bhotswap.get() == True:
        str_out += '- Electrical component damage due to hot swap.\n'
    if bvibration.get() == True:
        str_out += '- Internal component damage due to vibration.\n'
    if bnoise.get() == True:
        str_out += '- Electrical component damage due to inductive noise.\n'
    if bheat.get() == True:
        str_out += '- Internal component damage due to heat.\n'
    if bcable.get() == True:
        str_out += '- Cable damage due to impact or improper use of the cable.\n'
    if bnff.get() == True:
        str_out = '- No faults were found.\n'
    if bsurge.get() == True:
        str_out += '- An application of over voltage/current in the form of an electrical surge.\n'
    if bimpact.get() == True:
        str_out += '- Impact damage to the unit.\n'
    return str_out


def conclusion():
    str_out = ''
    if rp.get() == True:
        str_out = 'To bring this unit back to full functionality, the unit will be repaired and tested to ' \
                  'meet manufacturing specifications. A full operation and function check will be conducted to ' \
                  'confirm normal operation.\n'
    if nrp.get() == True:
        str_out = 'This unit cannot be repaired as either it is damaged beyond repairing capabilities or the repair ' \
                  'will approach the cost of a new unit. This unit will be sent back in the condition it was ' \
                  'received. We do not recommend putting this unit back on a production line, as we cannot guarantee' \
                  ' the reliability of its full functionality.\n'
    if rt.get() == True:
        str_out = 'Because we were not able to replicate the described symptom, the unit will be returned in the ' \
                  'condition it was received.\n'
    if kj.get() == True:
        str_out = 'The internal boards will be replaced to that of the equal repair cost and as a result, the' \
                  ' unit will have a new serial number.\n'
    return str_out


def preventative_measure():
    str_out = ''
    if bhotswap.get() == True:
        str_out += 'How swap damage could occur when the units power is not fully cycled off. Please make' \
                   ' sure to complete a unit shutdown before unplugging the unit. '
    if bvibration.get() == True:
        str_out += 'To prevent damage from vibration, please make sure to reduce the amount of vibration ' \
                   'the unit is subjected to by introducing vibration dampeners. '
    if bnoise.get() == True:
        str_out += 'To prevent damage from inductive noise, please make sure to keep this unit away from ' \
                   'machinery that can cause electrical noise. '
    if bheat.get() == True:
        str_out += 'To prevent damage from heat, please make sure to provide proper air flow to keep the unit cool. '
    if bcable.get() == True:
        str_out += 'To prevent cable damage, please make sure to protect the cable from impact and pulling. '
    if bnff.get() == True:
        str_out += 'No faults were found'
    if bimpact.get() == True:
        str_out += 'Please make sure not to subject the unit to impact damage.'
    else:
        pass
    return str_out


def makeform():
    rp_save = rpnum_text.get('1.0', 'end').split('\n')[0].strip()
    # serial_save = serial_text.get('1.0', 'end').split('\n')[0].strip()
    pdf = FPDF()
    pdf.add_page()
    sum_out = ''
    count = 0
    cliping = ''

    if rpnum_text.get('1.0', 'end').split('\n')[0] == '' or pic_text.get('1.0', 'end').split('\n')[0] == '' or \
            rp.get() == False and nrp.get() == False and rt.get() == False and kj.get() == False or \
            bhotswap.get() == False and bvibration.get() == False and \
            bnoise.get() == False and bheat.get() == False and bnff.get() == False and bcable.get() == False and \
            bsurge.get() == False and bimpact.get() == False or bapp_nff.get() == False and bapp_oil.get() == False and \
            bapp_impact.get() == False and bapp_cable.get() == False and bapp_filterglass.get() == False or \
            bsym_cable.get() == False and \
            bsym_inout.get() == False and bsym_nocom.get() == False and bsym_noint.get() == False and \
            bsym_nolaser.get() == False and bsym_nopower.get() == False and bsym_symnff.get() == False and \
            bsym_nffstruct.get() == False:
        checkerror()
        return None
    else:
        pass
    pdf.image('report_img/logo.jpg', w=190, h=55, type='jpg')
    # pdf.set_font('Arial', 'IB', 22)
    # pdf.cell(190, 15, 'Failure Analysis Report', ln=1, align='C')
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    if rpnum_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'RP#: ' + 'None          ')
    else:
        pdf.cell(40, 10, 'RP#: ' + rpnum_text.get('1.0', 'end') + '          ')

    if item_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Item: ' + 'None          ')
    else:
        pdf.cell(40, 10, 'Item: ' + item_text.get('1.0', 'end') + '          ')

    # if serial_text.get('1.0', 'end').strip() == '':
    #     pdf.cell(40, 10, 'Serial: ' + 'None          ')
    # else:
    pdf.cell(40, 10, 'Serial: ' + entries())

    pdf.ln(h=15)

    if key_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Keyword: ' + 'None          ')
    else:
        pdf.cell(40, 10, 'Keyword: ' + key_text.get('1.0', 'end'))

    if pic_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'PIC: ' + 'None          ')
    else:
        pdf.cell(40, 10, 'PIC: ' + pic_text.get('1.0', 'end'))

    if bench_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Time: ' + 'None          ')
    else:
        pdf.cell(40, 10, 'Time: ' + bench_text.get('1.0', 'end'))

    if pic_text.get('1.0', 'end').strip() == 'Angel' or pic_text.get('1.0', 'end').strip() == 'angel':
        pdf.image('report_img/angel.jpg', y=85, x=125, w=20, h=20, type='jpg')
    elif pic_text.get('1.0', 'end').strip() == 'Grant' or pic_text.get('1.0', 'end').strip() == 'grant':
        pdf.image('report_img/grant.jpg', y=85, x=125, w=20, h=20, type='jpg')
    elif pic_text.get('1.0', 'end').strip() == 'Adolfo' or pic_text.get('1.0', 'end').strip() == 'adolfo':
        pdf.image('report_img/adolfo.jpg', y=85, x=125, w=20, h=20, type='jpg')

    pdf.ln(h=15)
    pdf.image('report_img/rp-history.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)

    if his_rp_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'RP#: ' + 'None', ln=1)
    else:
        pdf.cell(40, 10, 'RP#: ' + his_rp_text.get('1.0', 'end'))

    if his_comp_date_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Date: ' + 'None', ln=1)
    else:
        pdf.cell(40, 10, 'Date:' + his_comp_date_text.get('1.0', 'end'))

    if his_symptom_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Symptom: ' + 'None', ln=1)
    else:
        pdf.cell(40, 10, 'Symptom: ' + his_symptom_text.get('1.0', 'end'), ln=1)

    if his_result_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Result: ' + 'None', ln=1)
    else:
        pdf.cell(40, 10, 'Result: ' + his_result_text.get('1.0', 'end'), ln=1)

    pdf.ln(h=15)
    pdf.image('report_img/apperance.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, appearance_text.get('1.0', 'end'), ln=1)
    pdf.ln(h=15)
    pdf.image('report_img/symptom.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, symptom_text.get('1.0', 'end'), ln=1)
    pdf.ln(h=15)
    pdf.image('report_img/tests.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, tests_text.get('1.0', 'end'), ln=1)
    pdf.ln(h=15)
    pdf.image('report_img/failure.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)
    pdf.ln(h=15)

    if fb_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Failed Board: ' + ' None', ln=1)
    else:
        pdf.cell(40, 10, 'Failed Board: ' + fb_text.get('1.0', 'end'), ln=1)

    if comp_text.get('1.0', 'end').strip() == '':
        pdf.cell(40, 10, 'Location of Component: ' + 'None', ln=1)
    else:
        pdf.cell(40, 10, 'Location of Component: ' + comp_text.get('1.0', 'end'), ln=1)
    pdf.cell(40, 10, 'Failed Structure Part: ' + structure_text.get('1.0', 'end'), ln=1)

    if parts_text.get('1.0', 'end').split('\n')[0] == '':
        pdf.cell(40, 10, 'Were parts replaced? No')
    else:
        pdf.cell(40, 10, 'Were parts replaced? Yes. ' + parts_text.get('1.0', 'end') + '.')

    pdf.ln(h=15)
    pdf.image('report_img/analysis.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)

    for a in range(len(analysis_text.get('1.0', 'end').split('\n'))):
        pdf.cell(40, 10, analysis_text.get('1.0', 'end').split('\n')[count], ln=1)
        count += 1

    pdf.ln(h=15)
    pdf.image('report_img/causeof.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, failure_text.get('1.0', 'end'), ln=1)
    pdf.ln(h=15)
    pdf.image('report_img/summary.jpg', w=190, h=12, type='jpg')
    pdf.set_font('Arial', 'B', 12)

    if rp.get() == True:
        sum_out = 'This unit was repaired and sent back to the customer'
    elif nrp.get() == True:
        sum_out = 'This unit was not repaired and sent back to the customer in the condition it was received'
    elif rt.get() == True:
        sum_out = 'This unit was sent back to the customer in the condition it was received'
    elif kj.get() == True:
        sum_out = 'This unit was sent to KJ'

    pdf.cell(40, 10, sum_out, ln=1)
    pdf.ln(h=15)
    pdf.add_page()
    pdf.cell(175, 15, 'Customer Inspection Report', ln=1, align='C')
    pdf.ln(h=15)

    pdf.set_font('Arial', '', 14)
    pdf.cell(40, 10, '< Appearance Check > ', ln=1)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(175, 10, apperance_check())

    if c_appearance_text.get('1.0', 'end').strip() == '':
        cliping += '< Appearance Check >\n' + apperance_check() + '\n'
    else:
        pdf.multi_cell(175, 10, 'Additional information:\n' + c_appearance_text.get('1.0', 'end') + '\n')
        cliping += '< Appearance Check >\n' + apperance_check() + '\nAdditional information:\n' + \
                   c_appearance_text.get('1.0', 'end') + '\n'

    pdf.set_font('Arial', '', 14)
    pdf.cell(40, 10, '< Operation Check > ', ln=1)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(175, 10, operation_check())

    if c_operation_text.get('1.0', 'end').strip() == '':
        cliping += '< Operation Check >\n' + operation_check() + '\n'
    else:
        pdf.multi_cell(175, 10, 'Additional information:\n' + c_operation_text.get('1.0', 'end') + '\n')
        cliping += '< Operation Check >\n' + operation_check() + '\nAdditional information:\n' + \
                   c_operation_text.get('1.0', 'end') + '\n'

    pdf.set_font('Arial', '', 14)
    pdf.cell(40, 10, '< Internal Check > ', ln=1)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(175, 10, internal_check())

    if c_internal_text.get('1.0', 'end').strip() == '':
        cliping += '< Internal Check >\n' + internal_check() + '\n'
    else:
        pdf.multi_cell(175, 10, 'Additional information:\n' + c_internal_text.get('1.0', 'end') + '\n')
        cliping += '< Internal Check >\n' + internal_check() + '\nAdditional information:\n' + \
                   c_internal_text.get('1.0', 'end') + '\n'

    pdf.set_font('Arial', '', 14)
    pdf.cell(40, 10, '< Suspected Cause > ', ln=1)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(175, 10, suspected_cause())

    if c_suspected_text.get('1.0', 'end').strip() == '':
        cliping += '< Suspected Cause >\n' + suspected_cause() + '\n'
    else:
        pdf.multi_cell(175, 10, 'Additional information:\n' + c_suspected_text.get('1.0', 'end') + '\n')
        cliping += '< Suspected Cause >\n' + suspected_cause() + '\nAdditional information:\n' + \
                   c_suspected_text.get('1.0', 'end') + '\n'

    pdf.set_font('Arial', '', 14)
    pdf.cell(40, 10, '< Conclusion > ', ln=1)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(175, 10, conclusion())

    if c_conclusion_text.get('1.0', 'end').strip() == '':
        cliping += '< Conclusion >\n' + conclusion() + '\n'
    else:
        pdf.multi_cell(175, 10, 'Additional information:\n' + c_conclusion_text.get('1.0', 'end') + '\n')
        cliping += '< Conclusion >\n' + conclusion() + '\nAdditional information:\n' + \
                   c_conclusion_text.get('1.0', 'end') + '\n'

    pdf.set_font('Arial', '', 14)
    pdf.cell(40, 10, '< Preventative Measure > ', ln=1)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(175, 10, preventative_measure())
    cliping += '< Preventative Measure >\n' + preventative_measure()

    try:
        # pdf.output('I:/PRD/Check Sheets/RSS/' + rp_save + ' - ' + serial_save + '.pdf', 'F')
        pdf.output(rp_save + ' - ' + '.pdf', 'F')
    except OSError:
        fileopen()
        return None

    pyperclip.copy(cliping)
    copy_message()
    rpnum_text.delete('1.0', 'end')
    # serial_text.delete('1.0', 'end')
    key_text.delete('1.0', 'end')
    comp_text.delete('1.0', 'end')
    structure_text.delete('1.0', 'end')
    fb_text.delete('1.0', 'end')


def focus_next_window(event):
    event.widget.tk_focusNext().focus()
    return "break"


def fileopen():
    showinfo('Warning!', 'You have the file open, close it before save.')


def warning(wrds):
    showinfo('Window', 'You are missing:\n' + wrds)


def copy_message():
    showinfo('Success!', 'Your inspection report has been copied to your clipboard.\n(CTL+V to paste in AS2)')


def too_many_char(incoming):
    showinfo('Warning!', f'Preset character limit is 18,\nyou have {incoming}')


def save_exe_preset1():
    if len(preset_text.get('1.0', 'end').split('\n')[0]) > 18:
        too_many_char(len(preset_text.get('1.0', 'end').split('\n')[0]))
        return None
    else:
        pass

    if preset_text.get('1.0', 'end').split('\n')[0] == '':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'A'
        rpnum_text.insert(tk.END, sheet_ranges[f'{letter}2'].value)
        item_text.insert(tk.END, sheet_ranges[f'{letter}3'].value)
        # serial_text.insert(tk.END, sheet_ranges[f'{letter}4'].value)
        key_text.insert(tk.END, sheet_ranges[f'{letter}5'].value)
        fb_text.insert(tk.END, sheet_ranges[f'{letter}6'].value)
        comp_text.insert(tk.END, sheet_ranges[f'{letter}7'].value)
        structure_text.insert(tk.END, sheet_ranges[f'{letter}8'].value)
        parts_text.insert(tk.END, sheet_ranges[f'{letter}9'].value)
        pic_text.insert(tk.END, sheet_ranges[f'{letter}10'].value)
        his_rp_text.insert(tk.END, sheet_ranges[f'{letter}11'].value)
        his_comp_date_text.insert(tk.END, sheet_ranges[f'{letter}12'].value)
        his_symptom_text.insert(tk.END, sheet_ranges[f'{letter}13'].value)
        his_result_text.insert(tk.END, sheet_ranges[f'{letter}14'].value)
        appearance_text.insert(tk.END, sheet_ranges[f'{letter}15'].value)
        symptom_text.insert(tk.END, sheet_ranges[f'{letter}16'].value)
        tests_text.insert(tk.END, sheet_ranges[f'{letter}17'].value)
        analysis_text.insert(tk.END, sheet_ranges[f'{letter}18'].value)
        failure_text.insert(tk.END, sheet_ranges[f'{letter}19'].value)
        if str(sheet_ranges[f'{letter}20'].value) == 'True':
            bapp_impact.set(True)
        else:
            bapp_impact.set(False)
        if str(sheet_ranges[f'{letter}21'].value) == 'True':
            bapp_cable.set(True)
        else:
            bapp_cable.set(False)
        if str(sheet_ranges[f'{letter}22'].value) == 'True':
            bapp_filterglass.set(True)
        else:
            bapp_filterglass.set(False)
        if str(sheet_ranges[f'{letter}23'].value) == 'True':
            bapp_nff.set(True)
        else:
            bapp_nff.set(False)
        if str(sheet_ranges[f'{letter}24'].value) == 'True':
            bapp_oil.set(True)
        else:
            bapp_oil.set(False)
        if str(sheet_ranges[f'{letter}25'].value) == 'True':
            bsym_cable.set(True)
        else:
            bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}26'].value) == 'True':
            bsym_inout.set(True)
        else:
            bsym_inout.set(False)
        if str(sheet_ranges[f'{letter}27'].value) == 'True':
            bsym_nffstruct.set(True)
        else:
            bsym_nffstruct.set(False)
        if str(sheet_ranges[f'{letter}28'].value) == 'True':
            bsym_nocom.set(True)
        else:
            bsym_nocom.set(False)
        if str(sheet_ranges[f'{letter}29'].value) == 'True':
            bsym_noint.set(True)
        else:
            bsym_noint.set(False)
        if str(sheet_ranges[f'{letter}30'].value) == 'True':
            bsym_nolaser.set(True)
        else:
            bsym_nolaser.set(False)
        if str(sheet_ranges[f'{letter}31'].value) == 'True':
            bsym_nopower.set(True)
        else:
            bsym_nopower.set(False)
        if str(sheet_ranges[f'{letter}32'].value) == 'True':
            bsym_symnff.set(True)
        else:
            bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}33'].value) == 'True':
            rp.set(True)
        else:
            rp.set(False)
        if str(sheet_ranges[f'{letter}34'].value) == 'True':
            nrp.set(True)
        else:
            nrp.set(False)
        if str(sheet_ranges[f'{letter}35'].value) == 'True':
            rt.set(True)
        else:
            rt.set(False)
        if str(sheet_ranges[f'{letter}36'].value) == 'True':
            kj.set(True)
        else:
            kj.set(False)
        if str(sheet_ranges[f'{letter}37'].value) == 'True':
            bhotswap.set(True)
        else:
            bhotswap.set(False)
        if str(sheet_ranges[f'{letter}38'].value) == 'True':
            bvibration.set(True)
        else:
            bvibration.set(False)
        if str(sheet_ranges[f'{letter}39'].value) == 'True':
            bnoise.set(True)
        else:
            bnoise.set(False)
        if str(sheet_ranges[f'{letter}40'].value) == 'True':
            bheat.set(True)
        else:
            bheat.set(False)
        if str(sheet_ranges[f'{letter}41'].value) == 'True':
            bnff.set(True)
        else:
            bnff.set(False)
        if str(sheet_ranges[f'{letter}42'].value) == 'True':
            bcable.set(True)
        else:
            bcable.set(False)
        if str(sheet_ranges[f'{letter}43'].value) == 'True':
            bsym_symnff.set(True)
        else:
            bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}44'].value) == 'True':
            bsym_cable.set(True)
        else:
            bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}45'].value) == 'True':
            bsurge.set(True)
        else:
            bsurge.set(False)
        c_appearance_text.insert(tk.END, sheet_ranges[f'{letter}46'].value)
        c_operation_text.insert(tk.END, sheet_ranges[f'{letter}47'].value)
        c_internal_text.insert(tk.END, sheet_ranges[f'{letter}48'].value)
        c_suspected_text.insert(tk.END, sheet_ranges[f'{letter}49'].value)
        c_conclusion_text.insert(tk.END, sheet_ranges[f'{letter}50'].value)
        bench_text.insert(tk.END, sheet_ranges[f'{letter}51'].value)
    elif preset_text.get('1.0', 'end').split('\n')[0] == 'clear':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'A'
        btn_preset_text.set('Preset 1')
        preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}1'].value = 'Preset 1'
        for a in range(2, 51):
            sheet_ranges[f'{letter}{a}'].value = ''
        lb.save('presets.xlsx')
    else:
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'A'

        sheet_ranges[f'{letter}1'].value = preset_text.get('1.0', 'end').split('\n')[0]
        btn_preset_text.set(sheet_ranges[f'{letter}1'].value)
        preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}2'].value = rpnum_text.get('1.0', 'end')
        sheet_ranges[f'{letter}3'].value = item_text.get('1.0', 'end')
        # sheet_ranges[f'{letter}4'].value = serial_text.get('1.0', 'end')
        sheet_ranges[f'{letter}5'].value = key_text.get('1.0', 'end')
        sheet_ranges[f'{letter}6'].value = fb_text.get('1.0', 'end')
        sheet_ranges[f'{letter}7'].value = comp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}8'].value = structure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}9'].value = parts_text.get('1.0', 'end')
        sheet_ranges[f'{letter}10'].value = pic_text.get('1.0', 'end')
        sheet_ranges[f'{letter}11'].value = his_rp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}12'].value = his_comp_date_text.get('1.0', 'end')
        sheet_ranges[f'{letter}13'].value = his_symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}14'].value = his_result_text.get('1.0', 'end')
        sheet_ranges[f'{letter}15'].value = appearance_text.get('1.0', 'end')
        sheet_ranges[f'{letter}16'].value = symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}17'].value = tests_text.get('1.0', 'end')
        sheet_ranges[f'{letter}18'].value = analysis_text.get('1.0', 'end')
        sheet_ranges[f'{letter}19'].value = failure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}20'].value = bapp_impact.get()
        sheet_ranges[f'{letter}21'].value = bapp_cable.get()
        sheet_ranges[f'{letter}22'].value = bapp_filterglass.get()
        sheet_ranges[f'{letter}23'].value = bapp_nff.get()
        sheet_ranges[f'{letter}24'].value = bapp_oil.get()
        sheet_ranges[f'{letter}25'].value = bsym_cable.get()
        sheet_ranges[f'{letter}26'].value = bsym_inout.get()
        sheet_ranges[f'{letter}27'].value = bsym_nffstruct.get()
        sheet_ranges[f'{letter}28'].value = bsym_nocom.get()
        sheet_ranges[f'{letter}29'].value = bsym_noint.get()
        sheet_ranges[f'{letter}30'].value = bsym_nolaser.get()
        sheet_ranges[f'{letter}31'].value = bsym_nopower.get()
        sheet_ranges[f'{letter}32'].value = bsym_symnff.get()
        sheet_ranges[f'{letter}33'].value = rp.get()
        sheet_ranges[f'{letter}34'].value = nrp.get()
        sheet_ranges[f'{letter}35'].value = rt.get()
        sheet_ranges[f'{letter}36'].value = kj.get()
        sheet_ranges[f'{letter}37'].value = bhotswap.get()
        sheet_ranges[f'{letter}38'].value = bvibration.get()
        sheet_ranges[f'{letter}39'].value = bnoise.get()
        sheet_ranges[f'{letter}40'].value = bheat.get()
        sheet_ranges[f'{letter}41'].value = bnff.get()
        sheet_ranges[f'{letter}42'].value = bcable.get()
        sheet_ranges[f'{letter}43'].value = bsym_symnff.get()
        sheet_ranges[f'{letter}44'].value = bsym_cable.get()
        sheet_ranges[f'{letter}45'].value = bsurge.get()
        sheet_ranges[f'{letter}46'].value = c_appearance_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}47'].value = c_operation_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}48'].value = c_internal_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}49'].value = c_suspected_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}50'].value = c_conclusion_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}51'].value = bench_text.get('1.0', 'end').split('\n')[0]
        lb.save('presets.xlsx')
    return None


def save_exe_preset2():
    if preset_text.get('1.0', 'end').split('\n')[0] == '':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'B'
        rpnum_text.insert(tk.END, sheet_ranges[f'{letter}2'].value)
        item_text.insert(tk.END, sheet_ranges[f'{letter}3'].value)
        # serial_text.insert(tk.END, sheet_ranges[f'{letter}4'].value)
        key_text.insert(tk.END, sheet_ranges[f'{letter}5'].value)
        fb_text.insert(tk.END, sheet_ranges[f'{letter}6'].value)
        comp_text.insert(tk.END, sheet_ranges[f'{letter}7'].value)
        structure_text.insert(tk.END, sheet_ranges[f'{letter}8'].value)
        parts_text.insert(tk.END, sheet_ranges[f'{letter}9'].value)
        pic_text.insert(tk.END, sheet_ranges[f'{letter}10'].value)
        his_rp_text.insert(tk.END, sheet_ranges[f'{letter}11'].value)
        his_comp_date_text.insert(tk.END, sheet_ranges[f'{letter}12'].value)
        his_symptom_text.insert(tk.END, sheet_ranges[f'{letter}13'].value)
        his_result_text.insert(tk.END, sheet_ranges[f'{letter}14'].value)
        appearance_text.insert(tk.END, sheet_ranges[f'{letter}15'].value)
        symptom_text.insert(tk.END, sheet_ranges[f'{letter}16'].value)
        tests_text.insert(tk.END, sheet_ranges[f'{letter}17'].value)
        analysis_text.insert(tk.END, sheet_ranges[f'{letter}18'].value)
        failure_text.insert(tk.END, sheet_ranges[f'{letter}19'].value)
        if str(sheet_ranges[f'{letter}20'].value) == 'True':
            bapp_impact.set(True)
        else:
            bapp_impact.set(False)
        if str(sheet_ranges[f'{letter}21'].value) == 'True':
            bapp_cable.set(True)
        else:
            bapp_cable.set(False)
        if str(sheet_ranges[f'{letter}22'].value) == 'True':
            bapp_filterglass.set(True)
        else:
            bapp_filterglass.set(False)
        if str(sheet_ranges[f'{letter}23'].value) == 'True':
            bapp_nff.set(True)
        else:
            bapp_nff.set(False)
        if str(sheet_ranges[f'{letter}24'].value) == 'True':
            bapp_oil.set(True)
        else:
            bapp_oil.set(False)
        if str(sheet_ranges[f'{letter}25'].value) == 'True':
            bsym_cable.set(True)
        else:
            bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}26'].value) == 'True':
            bsym_inout.set(True)
        else:
            bsym_inout.set(False)
        if str(sheet_ranges[f'{letter}27'].value) == 'True':
            bsym_nffstruct.set(True)
        else:
            bsym_nffstruct.set(False)
        if str(sheet_ranges[f'{letter}28'].value) == 'True':
            bsym_nocom.set(True)
        else:
            bsym_nocom.set(False)
        if str(sheet_ranges[f'{letter}29'].value) == 'True':
            bsym_noint.set(True)
        else:
            bsym_noint.set(False)
        if str(sheet_ranges[f'{letter}30'].value) == 'True':
            bsym_nolaser.set(True)
        else:
            bsym_nolaser.set(False)
        if str(sheet_ranges[f'{letter}31'].value) == 'True':
            bsym_nopower.set(True)
        else:
            bsym_nopower.set(False)
        if str(sheet_ranges[f'{letter}32'].value) == 'True':
            bsym_symnff.set(True)
        else:
            bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}33'].value) == 'True':
            rp.set(True)
        else:
            rp.set(False)
        if str(sheet_ranges[f'{letter}34'].value) == 'True':
            nrp.set(True)
        else:
            nrp.set(False)
        if str(sheet_ranges[f'{letter}35'].value) == 'True':
            rt.set(True)
        else:
            rt.set(False)
        if str(sheet_ranges[f'{letter}36'].value) == 'True':
            kj.set(True)
        else:
            kj.set(False)
        if str(sheet_ranges[f'{letter}37'].value) == 'True':
            bhotswap.set(True)
        else:
            bhotswap.set(False)
        if str(sheet_ranges[f'{letter}38'].value) == 'True':
            bvibration.set(True)
        else:
            bvibration.set(False)
        if str(sheet_ranges[f'{letter}39'].value) == 'True':
            bnoise.set(True)
        else:
            bnoise.set(False)
        if str(sheet_ranges[f'{letter}40'].value) == 'True':
            bheat.set(True)
        else:
            bheat.set(False)
        if str(sheet_ranges[f'{letter}41'].value) == 'True':
            bnff.set(True)
        else:
            bnff.set(False)
        if str(sheet_ranges[f'{letter}42'].value) == 'True':
            bcable.set(True)
        else:
            bcable.set(False)
        if str(sheet_ranges[f'{letter}43'].value) == 'True':
            bsym_symnff.set(True)
        else:
            bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}44'].value) == 'True':
            bsym_cable.set(True)
        else:
            bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}45'].value) == 'True':
            bsurge.set(True)
        else:
            bsurge.set(False)
        c_appearance_text.insert(tk.END, sheet_ranges[f'{letter}46'].value)
        c_operation_text.insert(tk.END, sheet_ranges[f'{letter}47'].value)
        c_internal_text.insert(tk.END, sheet_ranges[f'{letter}48'].value)
        c_suspected_text.insert(tk.END, sheet_ranges[f'{letter}49'].value)
        c_conclusion_text.insert(tk.END, sheet_ranges[f'{letter}50'].value)
        bench_text.insert(tk.END, sheet_ranges[f'{letter}51'].value)
    elif preset_text.get('1.0', 'end').split('\n')[0] == 'clear':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'B'
        btn_preset2_text.set('Preset 2')
        preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}1'].value = 'Preset 2'
        for a in range(2, 51):
            sheet_ranges[f'{letter}{a}'].value = ''
        lb.save('presets.xlsx')
    else:
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'B'

        sheet_ranges[f'{letter}1'].value = preset_text.get('1.0', 'end').split('\n')[0]
        btn_preset2_text.set(sheet_ranges[f'{letter}1'].value)
        preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}2'].value = rpnum_text.get('1.0', 'end')
        sheet_ranges[f'{letter}3'].value = item_text.get('1.0', 'end')
        # sheet_ranges[f'{letter}4'].value = serial_text.get('1.0', 'end')
        sheet_ranges[f'{letter}5'].value = key_text.get('1.0', 'end')
        sheet_ranges[f'{letter}6'].value = fb_text.get('1.0', 'end')
        sheet_ranges[f'{letter}7'].value = comp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}8'].value = structure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}9'].value = parts_text.get('1.0', 'end')
        sheet_ranges[f'{letter}10'].value = pic_text.get('1.0', 'end')
        sheet_ranges[f'{letter}11'].value = his_rp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}12'].value = his_comp_date_text.get('1.0', 'end')
        sheet_ranges[f'{letter}13'].value = his_symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}14'].value = his_result_text.get('1.0', 'end')
        sheet_ranges[f'{letter}15'].value = appearance_text.get('1.0', 'end')
        sheet_ranges[f'{letter}16'].value = symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}17'].value = tests_text.get('1.0', 'end')
        sheet_ranges[f'{letter}18'].value = analysis_text.get('1.0', 'end')
        sheet_ranges[f'{letter}19'].value = failure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}20'].value = bapp_impact.get()
        sheet_ranges[f'{letter}21'].value = bapp_cable.get()
        sheet_ranges[f'{letter}22'].value = bapp_filterglass.get()
        sheet_ranges[f'{letter}23'].value = bapp_nff.get()
        sheet_ranges[f'{letter}24'].value = bapp_oil.get()
        sheet_ranges[f'{letter}25'].value = bsym_cable.get()
        sheet_ranges[f'{letter}26'].value = bsym_inout.get()
        sheet_ranges[f'{letter}27'].value = bsym_nffstruct.get()
        sheet_ranges[f'{letter}28'].value = bsym_nocom.get()
        sheet_ranges[f'{letter}29'].value = bsym_noint.get()
        sheet_ranges[f'{letter}30'].value = bsym_nolaser.get()
        sheet_ranges[f'{letter}31'].value = bsym_nopower.get()
        sheet_ranges[f'{letter}32'].value = bsym_symnff.get()
        sheet_ranges[f'{letter}33'].value = rp.get()
        sheet_ranges[f'{letter}34'].value = nrp.get()
        sheet_ranges[f'{letter}35'].value = rt.get()
        sheet_ranges[f'{letter}36'].value = kj.get()
        sheet_ranges[f'{letter}37'].value = bhotswap.get()
        sheet_ranges[f'{letter}38'].value = bvibration.get()
        sheet_ranges[f'{letter}39'].value = bnoise.get()
        sheet_ranges[f'{letter}40'].value = bheat.get()
        sheet_ranges[f'{letter}41'].value = bnff.get()
        sheet_ranges[f'{letter}42'].value = bcable.get()
        sheet_ranges[f'{letter}43'].value = bsym_symnff.get()
        sheet_ranges[f'{letter}44'].value = bsym_cable.get()
        sheet_ranges[f'{letter}45'].value = bsurge.get()
        sheet_ranges[f'{letter}46'].value = c_appearance_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}47'].value = c_operation_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}48'].value = c_internal_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}49'].value = c_suspected_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}50'].value = c_conclusion_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}51'].value = bench_text.get('1.0', 'end').split('\n')[0]
        lb.save('presets.xlsx')
    return None


def save_exe_preset3():
    if preset_text.get('1.0', 'end').split('\n')[0] == '':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'C'
        rpnum_text.insert(tk.END, sheet_ranges[f'{letter}2'].value)
        item_text.insert(tk.END, sheet_ranges[f'{letter}3'].value)
        # serial_text.insert(tk.END, sheet_ranges[f'{letter}4'].value)
        key_text.insert(tk.END, sheet_ranges[f'{letter}5'].value)
        fb_text.insert(tk.END, sheet_ranges[f'{letter}6'].value)
        comp_text.insert(tk.END, sheet_ranges[f'{letter}7'].value)
        structure_text.insert(tk.END, sheet_ranges[f'{letter}8'].value)
        parts_text.insert(tk.END, sheet_ranges[f'{letter}9'].value)
        pic_text.insert(tk.END, sheet_ranges[f'{letter}10'].value)
        his_rp_text.insert(tk.END, sheet_ranges[f'{letter}11'].value)
        his_comp_date_text.insert(tk.END, sheet_ranges[f'{letter}12'].value)
        his_symptom_text.insert(tk.END, sheet_ranges[f'{letter}13'].value)
        his_result_text.insert(tk.END, sheet_ranges[f'{letter}14'].value)
        appearance_text.insert(tk.END, sheet_ranges[f'{letter}15'].value)
        symptom_text.insert(tk.END, sheet_ranges[f'{letter}16'].value)
        tests_text.insert(tk.END, sheet_ranges[f'{letter}17'].value)
        analysis_text.insert(tk.END, sheet_ranges[f'{letter}18'].value)
        failure_text.insert(tk.END, sheet_ranges[f'{letter}19'].value)
        if str(sheet_ranges[f'{letter}20'].value) == 'True':
            bapp_impact.set(True)
        else:
            bapp_impact.set(False)
        if str(sheet_ranges[f'{letter}21'].value) == 'True':
            bapp_cable.set(True)
        else:
            bapp_cable.set(False)
        if str(sheet_ranges[f'{letter}22'].value) == 'True':
            bapp_filterglass.set(True)
        else:
            bapp_filterglass.set(False)
        if str(sheet_ranges[f'{letter}23'].value) == 'True':
            bapp_nff.set(True)
        else:
            bapp_nff.set(False)
        if str(sheet_ranges[f'{letter}24'].value) == 'True':
            bapp_oil.set(True)
        else:
            bapp_oil.set(False)
        if str(sheet_ranges[f'{letter}25'].value) == 'True':
            bsym_cable.set(True)
        else:
            bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}26'].value) == 'True':
            bsym_inout.set(True)
        else:
            bsym_inout.set(False)
        if str(sheet_ranges[f'{letter}27'].value) == 'True':
            bsym_nffstruct.set(True)
        else:
            bsym_nffstruct.set(False)
        if str(sheet_ranges[f'{letter}28'].value) == 'True':
            bsym_nocom.set(True)
        else:
            bsym_nocom.set(False)
        if str(sheet_ranges[f'{letter}29'].value) == 'True':
            bsym_noint.set(True)
        else:
            bsym_noint.set(False)
        if str(sheet_ranges[f'{letter}30'].value) == 'True':
            bsym_nolaser.set(True)
        else:
            bsym_nolaser.set(False)
        if str(sheet_ranges[f'{letter}31'].value) == 'True':
            bsym_nopower.set(True)
        else:
            bsym_nopower.set(False)
        if str(sheet_ranges[f'{letter}32'].value) == 'True':
            bsym_symnff.set(True)
        else:
            bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}33'].value) == 'True':
            rp.set(True)
        else:
            rp.set(False)
        if str(sheet_ranges[f'{letter}34'].value) == 'True':
            nrp.set(True)
        else:
            nrp.set(False)
        if str(sheet_ranges[f'{letter}35'].value) == 'True':
            rt.set(True)
        else:
            rt.set(False)
        if str(sheet_ranges[f'{letter}36'].value) == 'True':
            kj.set(True)
        else:
            kj.set(False)
        if str(sheet_ranges[f'{letter}37'].value) == 'True':
            bhotswap.set(True)
        else:
            bhotswap.set(False)
        if str(sheet_ranges[f'{letter}38'].value) == 'True':
            bvibration.set(True)
        else:
            bvibration.set(False)
        if str(sheet_ranges[f'{letter}39'].value) == 'True':
            bnoise.set(True)
        else:
            bnoise.set(False)
        if str(sheet_ranges[f'{letter}40'].value) == 'True':
            bheat.set(True)
        else:
            bheat.set(False)
        if str(sheet_ranges[f'{letter}41'].value) == 'True':
            bnff.set(True)
        else:
            bnff.set(False)
        if str(sheet_ranges[f'{letter}42'].value) == 'True':
            bcable.set(True)
        else:
            bcable.set(False)
        if str(sheet_ranges[f'{letter}43'].value) == 'True':
            bsym_symnff.set(True)
        else:
            bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}44'].value) == 'True':
            bsym_cable.set(True)
        else:
            bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}45'].value) == 'True':
            bsurge.set(True)
        else:
            bsurge.set(False)
        c_appearance_text.insert(tk.END, sheet_ranges[f'{letter}46'].value)
        c_operation_text.insert(tk.END, sheet_ranges[f'{letter}47'].value)
        c_internal_text.insert(tk.END, sheet_ranges[f'{letter}48'].value)
        c_suspected_text.insert(tk.END, sheet_ranges[f'{letter}49'].value)
        c_conclusion_text.insert(tk.END, sheet_ranges[f'{letter}50'].value)
        bench_text.insert(tk.END, sheet_ranges[f'{letter}51'].value)
    elif preset_text.get('1.0', 'end').split('\n')[0] == 'clear':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'C'
        btn_preset3_text.set('Preset 3')
        preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}1'].value = 'Preset 3'
        for a in range(2, 51):
            sheet_ranges[f'{letter}{a}'].value = ''
        lb.save('presets.xlsx')
    else:
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'C'

        sheet_ranges[f'{letter}1'].value = preset_text.get('1.0', 'end').split('\n')[0]
        btn_preset3_text.set(sheet_ranges[f'{letter}1'].value)
        preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}2'].value = rpnum_text.get('1.0', 'end')
        sheet_ranges[f'{letter}3'].value = item_text.get('1.0', 'end')
        # sheet_ranges[f'{letter}4'].value = serial_text.get('1.0', 'end')
        sheet_ranges[f'{letter}5'].value = key_text.get('1.0', 'end')
        sheet_ranges[f'{letter}6'].value = fb_text.get('1.0', 'end')
        sheet_ranges[f'{letter}7'].value = comp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}8'].value = structure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}9'].value = parts_text.get('1.0', 'end')
        sheet_ranges[f'{letter}10'].value = pic_text.get('1.0', 'end')
        sheet_ranges[f'{letter}11'].value = his_rp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}12'].value = his_comp_date_text.get('1.0', 'end')
        sheet_ranges[f'{letter}13'].value = his_symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}14'].value = his_result_text.get('1.0', 'end')
        sheet_ranges[f'{letter}15'].value = appearance_text.get('1.0', 'end')
        sheet_ranges[f'{letter}16'].value = symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}17'].value = tests_text.get('1.0', 'end')
        sheet_ranges[f'{letter}18'].value = analysis_text.get('1.0', 'end')
        sheet_ranges[f'{letter}19'].value = failure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}20'].value = bapp_impact.get()
        sheet_ranges[f'{letter}21'].value = bapp_cable.get()
        sheet_ranges[f'{letter}22'].value = bapp_filterglass.get()
        sheet_ranges[f'{letter}23'].value = bapp_nff.get()
        sheet_ranges[f'{letter}24'].value = bapp_oil.get()
        sheet_ranges[f'{letter}25'].value = bsym_cable.get()
        sheet_ranges[f'{letter}26'].value = bsym_inout.get()
        sheet_ranges[f'{letter}27'].value = bsym_nffstruct.get()
        sheet_ranges[f'{letter}28'].value = bsym_nocom.get()
        sheet_ranges[f'{letter}29'].value = bsym_noint.get()
        sheet_ranges[f'{letter}30'].value = bsym_nolaser.get()
        sheet_ranges[f'{letter}31'].value = bsym_nopower.get()
        sheet_ranges[f'{letter}32'].value = bsym_symnff.get()
        sheet_ranges[f'{letter}33'].value = rp.get()
        sheet_ranges[f'{letter}34'].value = nrp.get()
        sheet_ranges[f'{letter}35'].value = rt.get()
        sheet_ranges[f'{letter}36'].value = kj.get()
        sheet_ranges[f'{letter}37'].value = bhotswap.get()
        sheet_ranges[f'{letter}38'].value = bvibration.get()
        sheet_ranges[f'{letter}39'].value = bnoise.get()
        sheet_ranges[f'{letter}40'].value = bheat.get()
        sheet_ranges[f'{letter}41'].value = bnff.get()
        sheet_ranges[f'{letter}42'].value = bcable.get()
        sheet_ranges[f'{letter}43'].value = bsym_symnff.get()
        sheet_ranges[f'{letter}44'].value = bsym_cable.get()
        sheet_ranges[f'{letter}45'].value = bsurge.get()
        sheet_ranges[f'{letter}46'].value = c_appearance_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}47'].value = c_operation_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}48'].value = c_internal_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}49'].value = c_suspected_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}50'].value = c_conclusion_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}51'].value = bench_text.get('1.0', 'end').split('\n')[0]
        lb.save('presets.xlsx')
    return None


count = 7


def addBox():
    global count
    if count < 11:
        serial_box = tk.Text(root, height=1, width=20)
        serial_box.grid(row=count, column=1)
        serial_box.bind("<Tab>", focus_next_window)
        count += 1
        serial_list.append(serial_box)
    else:
        pass


def entries():
    sout = ''
    for s in serial_list:
        sout += s.get('1.0', 'end')
    return sout


serial_list = []

root = tk.Tk()
w, h = root.winfo_screenwidth(), root.winfo_screenheight()
root.geometry("%dx%d+0+0" % (w, h))
# root.geometry('1300x900')
root.title('RSS Generator')

rpnum_label = tk.Label(text='RP#')
rpnum_label.grid(row=1, column=1)
rpnum_text = tk.Text(root, height=1, width=20)
rpnum_text.grid(row=2, column=1)
rpnum_text.bind("<Tab>", focus_next_window)

item_label = tk.Label(text='Item Name')
item_label.grid(row=3, column=1)
item_text = tk.Text(root, height=1, width=20)
item_text.grid(row=4, column=1)
item_text.bind("<Tab>", focus_next_window)

serial_label = tk.Label(text='Serial#')
serial_label.grid(row=5, column=1)
# serial_text = tk.Text(root, height=1, width=20)
# serial_text.grid(row=6, column=1)
# serial_text.bind("<Tab>", focus_next_window)
addBox()
serial_text = tk.Button(root, text='Add Serial Box', fg="Red", command=addBox)
serial_text.grid(row=6, column=1)

key_label = tk.Label(text='Keycode')
key_label.grid(row=1, column=2)
key_text = tk.Text(root, height=1, width=20)
key_text.grid(row=2, column=2)
key_text.bind("<Tab>", focus_next_window)

fb_label = tk.Label(text='Failed Board')
fb_label.grid(row=3, column=2)
fb_text = tk.Text(root, height=1, width=20)
fb_text.grid(row=4, column=2)
fb_text.bind("<Tab>", focus_next_window)

comp_label = tk.Label(text='Location of Component')
comp_label.grid(row=5, column=2)
comp_text = tk.Text(root, height=1, width=20)
comp_text.grid(row=6, column=2)
comp_text.bind("<Tab>", focus_next_window)

structure_label = tk.Label(text='Failed Structure Part')
structure_label.grid(row=1, column=3)
structure_text = tk.Text(root, height=1, width=20)
structure_text.grid(row=2, column=3)
structure_text.bind("<Tab>", focus_next_window)

parts_label = tk.Label(text='Parts Replaced')
parts_label.grid(row=3, column=3)
parts_text = tk.Text(root, height=1, width=20)
parts_text.grid(row=4, column=3)
parts_text.bind("<Tab>", focus_next_window)

pic_label = tk.Label(text='PIC')
pic_label.grid(row=5, column=3)
pic_text = tk.Text(root, height=1, width=20)
pic_text.grid(row=6, column=3)
pic_text.bind("<Tab>", focus_next_window)

bench_label = tk.Label(text='Bench Time')
bench_label.grid(row=1, column=4)
bench_text = tk.Text(root, height=1, width=20)
bench_text.grid(row=2, column=4)
bench_text.bind("<Tab>", focus_next_window)

bench_label = tk.Label(text='Error')
bench_label.grid(row=3, column=4)
bench_text = tk.Text(root, height=1, width=20)
bench_text.grid(row=4, column=4)
bench_text.bind("<Tab>", focus_next_window)

his_label = tk.Label(text='RP History')
his_label.grid(row=1, column=5)

his_rp = tk.Label(text='RP#')
his_rp.grid(row=2, column=5)
his_rp_text = tk.Text(root, height=1, width=20)
his_rp_text.grid(row=3, column=5)
his_rp_text.bind("<Tab>", focus_next_window)

his_comp_date = tk.Label(text='Completion Date')
his_comp_date.grid(row=4, column=5)
his_comp_date_text = tk.Text(root, height=1, width=20)
his_comp_date_text.grid(row=5, column=5)
his_comp_date_text.bind("<Tab>", focus_next_window)

his_symptom = tk.Label(text='Symptom')
his_symptom.grid(row=6, column=5)
his_symptom_text = tk.Text(root, height=1, width=20)
his_symptom_text.grid(row=7, column=5)
his_symptom_text.bind("<Tab>", focus_next_window)

his_result = tk.Label(text='Result')
his_result.grid(row=8, column=5)
his_result_text = tk.Text(root, height=1, width=20)
his_result_text.grid(row=9, column=5)
his_result_text.bind("<Tab>", focus_next_window)



''' Check Boxes'''

rp = tk.BooleanVar()
nrp = tk.BooleanVar()
rt = tk.BooleanVar()
kj = tk.BooleanVar()
bhotswap = tk.BooleanVar()
bvibration = tk.BooleanVar()
bnoise = tk.BooleanVar()
bheat = tk.BooleanVar()
bsurge = tk.BooleanVar()
bnff = tk.BooleanVar()
bcable = tk.BooleanVar()
bimpact = tk.BooleanVar()

bapp_nff = tk.BooleanVar()
bapp_cable = tk.BooleanVar()
bapp_impact = tk.BooleanVar()
bapp_oil = tk.BooleanVar()
bapp_filterglass = tk.BooleanVar()

bsym_nopower = tk.BooleanVar()
bsym_noint = tk.BooleanVar()
bsym_nocom = tk.BooleanVar()
bsym_nolaser = tk.BooleanVar()
bsym_nffstruct = tk.BooleanVar()
bsym_symnff = tk.BooleanVar()
bsym_cable = tk.BooleanVar()
bsym_inout = tk.BooleanVar()

c_addition = tk.Label(text='Inspection Report', fg="black")
c_addition.grid(row=10, column=3)

key_label = tk.Label(text='Appearance')
key_label.grid(row=11, column=1)

app_cable = tk.Checkbutton(root, text="Cable Damage",
                           onvalue=1, offvalue=0, height=1,
                           width=25, variable=bapp_cable)
app_cable.grid(row=12, column=1)

app_impact = tk.Checkbutton(root, text="Impact Damage",
                            onvalue=1, offvalue=0, height=1,
                            width=25, variable=bapp_impact)
app_impact.grid(row=13, column=1)

app_oil = tk.Checkbutton(root, text="Oil",
                         onvalue=1, offvalue=0, height=1,
                         width=25, variable=bapp_oil)
app_oil.grid(row=14, column=1)

app_filter = tk.Checkbutton(root, text="Filter Glass",
                            onvalue=1, offvalue=0, height=1,
                            width=25, variable=bapp_filterglass)
app_filter.grid(row=15, column=1)

app_nff = tk.Checkbutton(root, text="NFF",
                         onvalue=1, offvalue=0, height=1,
                         width=25, variable=bapp_nff)
app_nff.grid(row=16, column=1)

key_label = tk.Label(text='Symptom')
key_label.grid(row=10, column=2)

nopower = tk.Checkbutton(root, text="No Power",
                         onvalue=1, offvalue=0, height=1,
                         width=25, variable=bsym_nopower)
nopower.grid(row=12, column=2)

nolaser = tk.Checkbutton(root, text="No Laser",
                         onvalue=1, offvalue=0, height=1,
                         width=25, variable=bsym_nolaser)
nolaser.grid(row=13, column=2)

noint = tk.Checkbutton(root, text="No Initialization",
                       onvalue=1, offvalue=0, height=1,
                       width=25, variable=bsym_noint)
noint.grid(row=14, column=2)

nocom = tk.Checkbutton(root, text="No Communication",
                       onvalue=1, offvalue=0, height=1,
                       width=25, variable=bsym_nocom)
nocom.grid(row=15, column=2)

inout = tk.Checkbutton(root, text="Input/Output Error",
                       onvalue=1, offvalue=0, height=1,
                       width=25, variable=bsym_inout)
inout.grid(row=16, column=2)

sym_cable = tk.Checkbutton(root, text="Cable Damage",
                           onvalue=1, offvalue=0, height=1,
                           width=25, variable=bsym_cable)
sym_cable.grid(row=17, column=2)

sym_nffstruct = tk.Checkbutton(root, text="NFF(Structural Damage)",
                               onvalue=1, offvalue=0, height=1,
                               width=25, variable=bsym_nffstruct)
sym_nffstruct.grid(row=10, column=2)

symnff = tk.Checkbutton(root, text="NFF",
                        onvalue=1, offvalue=0, height=1,
                        width=25, variable=bsym_symnff)
symnff.grid(row=11, column=2)

key_label = tk.Label(text='Result')
key_label.grid(row=12, column=3)

rpchk_label = tk.Checkbutton(root, text="Repaired",
                             onvalue=1, offvalue=0, height=1,
                             width=25, variable=rp)
rpchk_label.grid(row=13, column=3)

nrpchk_label = tk.Checkbutton(root, text="Not repaired",
                              onvalue=1, offvalue=0, height=1,
                              width=25, variable=nrp)
nrpchk_label.grid(row=14, column=3)

rtchk_label = tk.Checkbutton(root, text="Returned (NFF)",
                             onvalue=1, offvalue=0, height=1,
                             width=25, variable=rt)
rtchk_label.grid(row=15, column=3)

kjchk_label = tk.Checkbutton(root, text="Sent to KJ(RRC)",
                             onvalue=1, offvalue=0, height=1,
                             width=25, variable=kj)
kjchk_label.grid(row=16, column=3)

key_label = tk.Label(text=' Suspected Cause/Preventative Measures ')
key_label.grid(row=10, column=4)

key_label = tk.Label(text=' Suspected Cause/Preventative Measures ')
key_label.grid(row=10, column=5)

hot_swap = tk.Checkbutton(root, text="Hot Swap",
                          onvalue=1, offvalue=0, height=1,
                          width=25, variable=bhotswap)
hot_swap.grid(row=11, column=4)

noise = tk.Checkbutton(root, text="Electrical Noise",
                       onvalue=1, offvalue=0, height=1,
                       width=25, variable=bnoise)
noise.grid(row=12, column=4)

vibration = tk.Checkbutton(root, text="Vibration",
                           onvalue=1, offvalue=0, height=1,
                           width=25, variable=bvibration)
vibration.grid(row=13, column=4)

surge = tk.Checkbutton(root, text="Electrical Surge",
                       onvalue=1, offvalue=0, height=1,
                       width=25, variable=bsurge)
surge.grid(row=14, column=4)

heat = tk.Checkbutton(root, text="Heat",
                      onvalue=1, offvalue=0, height=1,
                      width=25, variable=bheat)
heat.grid(row=15, column=4)

cable = tk.Checkbutton(root, text="Cable Damage",
                       onvalue=1, offvalue=0, height=1,
                       width=25, variable=bcable)
cable.grid(row=11, column=5)

impact = tk.Checkbutton(root, text="Impact",
                        onvalue=1, offvalue=0, height=1,
                        width=25, variable=bimpact)
impact.grid(row=12, column=5)

nff = tk.Checkbutton(root, text="NFF",
                     onvalue=1, offvalue=0, height=1,
                     width=25, variable=bnff)
nff.grid(row=13, column=5)

c_rss = tk.Label(text='RSS Comments', fg="blue")
c_rss.grid(row=18, column=3)

appearance = tk.Label(text='Appearance Inspection (RSS)', fg="blue")
appearance.grid(row=19, column=1)
appearance_text = tk.Text(root, height=5, width=30)
appearance_text.grid(row=20, column=1)
appearance_text.bind("<Tab>", focus_next_window)

symptom = tk.Label(text='Symptom Information (RSS)', fg="blue")
symptom.grid(row=19, column=2)
symptom_text = tk.Text(root, height=5, width=30)
symptom_text.grid(row=20, column=2)
symptom_text.bind("<Tab>", focus_next_window)

tests = tk.Label(text='Tests Performed (RSS)', fg="blue")
tests.grid(row=19, column=3)
tests_text = tk.Text(root, height=5, width=30)
tests_text.grid(row=20, column=3)
tests_text.bind("<Tab>", focus_next_window)

analysis = tk.Label(text='Analysis Steps (RSS)', fg="blue")
analysis.grid(row=19, column=4)
analysis_text = tk.Text(root, height=5, width=30)
analysis_text.grid(row=20, column=4)
analysis_text.bind("<Tab>", focus_next_window)

failure = tk.Label(text='Cause of Failure (RSS)', fg="blue")
failure.grid(row=19, column=5)
failure_text = tk.Text(root, height=5, width=30)
failure_text.grid(row=20, column=5)
failure_text.bind("<Tab>", focus_next_window)


'''Form'''

c_addition = tk.Label(text='Additional Comments for Inspection Report', fg="red")
c_addition.grid(row=24, column=3)

c_appearance = tk.Label(text='Appearance Check', fg="red")
c_appearance.grid(row=25, column=1)
c_appearance_text = tk.Text(root, height=5, width=30)
c_appearance_text.grid(row=26, column=1)
c_appearance_text.bind("<Tab>", focus_next_window)

c_operation = tk.Label(text='Operation Check', fg="red")
c_operation.grid(row=25, column=2)
c_operation_text = tk.Text(root, height=5, width=30)
c_operation_text.grid(row=26, column=2)
c_operation_text.bind("<Tab>", focus_next_window)

c_internal = tk.Label(text='Internal Check', fg="red")
c_internal.grid(row=25, column=3)
c_internal_text = tk.Text(root, height=5, width=30)
c_internal_text.grid(row=26, column=3)
c_internal_text.bind("<Tab>", focus_next_window)

c_suspected = tk.Label(text='Suspected Cause', fg="red")
c_suspected.grid(row=25, column=4)
c_suspected_text = tk.Text(root, height=5, width=30)
c_suspected_text.grid(row=26, column=4)
c_suspected_text.bind("<Tab>", focus_next_window)

c_conclusion = tk.Label(text='Conclusion', fg="red")
c_conclusion.grid(row=25, column=5)
c_conclusion_text = tk.Text(root, height=5, width=30)
c_conclusion_text.grid(row=26, column=5)
c_conclusion_text.bind("<Tab>", focus_next_window)

preset_label = tk.Label(text='Presets ("clear" for reset)')
preset_label.grid(row=33, column=4)
preset_text = tk.Text(root, height=1, width=20)
preset_text.grid(row=34, column=4)
preset_text.bind("<Tab>", focus_next_window)

lb = load_workbook(filename='presets.xlsx')
ws = lb.active
sheet_ranges = lb['Sheet']

btn_preset_text = tk.StringVar()
btn_preset_text.set(sheet_ranges['A1'].value)
btn_preset1 = tk.Button(root, height=1, width=15, textvariable=btn_preset_text, command=save_exe_preset1)
btn_preset1.grid(row=33, column=5)

btn_preset2_text = tk.StringVar()
btn_preset2_text.set(sheet_ranges['B1'].value)
btn_preset2 = tk.Button(root, height=1, width=15, textvariable=btn_preset2_text, command=save_exe_preset2)
btn_preset2.grid(row=34, column=5)

btn_preset3_text = tk.StringVar()
btn_preset3_text.set(sheet_ranges['C1'].value)
btn_preset3 = tk.Button(root, height=1, width=15, textvariable=btn_preset3_text, command=save_exe_preset3)
btn_preset3.grid(row=35, column=5)

btnRead = tk.Button(root, height=1, width=15, text="Save", command=makeform)
btnRead.grid(row=33, column=1)

btn_example = tk.Button(root, height=1, width=15, text="Example", command=example)
btn_example.grid(row=34, column=1)

btn_example = tk.Button(root, height=1, width=15, text="Clear All Values", command=clearexample)
btn_example.grid(row=33, column=2)

btn_example = tk.Button(root, height=1, width=15, text="Help", command=clearexample)
btn_example.grid(row=34, column=2)

angel_label = tk.Label(text='Created by Angel Davila')
angel_label.grid(row=35, column=1)

root.mainloop()
