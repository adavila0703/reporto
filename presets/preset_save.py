import gui.gui as gui
from openpyxl import load_workbook


def save_exe_preset1():
    """Save all values into a preset 1."""
    if len(gui.preset_text.get('1.0', 'end').split('\n')[0]) > 18:
        gui.too_many_char(len(gui.preset_text.get('1.0', 'end').split('\n')[0]))
        return None
    else:
        pass

    if gui.preset_text.get('1.0', 'end').split('\n')[0] == '':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'A'
        gui.rpnum_text.insert(gui.tk.END, sheet_ranges[f'{letter}2'].value)
        gui.item_text.insert(gui.tk.END, sheet_ranges[f'{letter}3'].value)
        # serial_text.insert(tk.END, sheet_ranges[f'{letter}4'].value)
        gui.key_text.insert(gui.tk.END, sheet_ranges[f'{letter}5'].value)
        gui.fb_text.insert(gui.tk.END, sheet_ranges[f'{letter}6'].value)
        gui.comp_text.insert(gui.tk.END, sheet_ranges[f'{letter}7'].value)
        gui.structure_text.insert(gui.tk.END, sheet_ranges[f'{letter}8'].value)
        gui.parts_text.insert(gui.tk.END, sheet_ranges[f'{letter}9'].value)
        gui.pic_text.insert(gui.tk.END, sheet_ranges[f'{letter}10'].value)
        gui.his_rp_text.insert(gui.tk.END, sheet_ranges[f'{letter}11'].value)
        gui.his_comp_date_text.insert(gui.tk.END, sheet_ranges[f'{letter}12'].value)
        gui.his_symptom_text.insert(gui.tk.END, sheet_ranges[f'{letter}13'].value)
        gui.his_result_text.insert(gui.tk.END, sheet_ranges[f'{letter}14'].value)
        gui.appearance_text.insert(gui.tk.END, sheet_ranges[f'{letter}15'].value)
        gui.symptom_text.insert(gui.tk.END, sheet_ranges[f'{letter}16'].value)
        gui.tests_text.insert(gui.tk.END, sheet_ranges[f'{letter}17'].value)
        gui.analysis_text.insert(gui.tk.END, sheet_ranges[f'{letter}18'].value)
        gui.failure_text.insert(gui.tk.END, sheet_ranges[f'{letter}19'].value)
        if str(sheet_ranges[f'{letter}20'].value) == 'True':
            gui.bapp_impact.set(True)
        else:
            gui.bapp_impact.set(False)
        if str(sheet_ranges[f'{letter}21'].value) == 'True':
            gui.bapp_cable.set(True)
        else:
            gui.bapp_cable.set(False)
        if str(sheet_ranges[f'{letter}22'].value) == 'True':
            gui.bapp_filterglass.set(True)
        else:
            gui.bapp_filterglass.set(False)
        if str(sheet_ranges[f'{letter}23'].value) == 'True':
            gui.bapp_nff.set(True)
        else:
            gui.bapp_nff.set(False)
        if str(sheet_ranges[f'{letter}24'].value) == 'True':
            gui.bapp_oil.set(True)
        else:
            gui.bapp_oil.set(False)
        if str(sheet_ranges[f'{letter}25'].value) == 'True':
            gui.bsym_cable.set(True)
        else:
            gui.bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}26'].value) == 'True':
            gui.bsym_inout.set(True)
        else:
            gui.bsym_inout.set(False)
        if str(sheet_ranges[f'{letter}27'].value) == 'True':
            gui.bsym_nffstruct.set(True)
        else:
            gui.bsym_nffstruct.set(False)
        if str(sheet_ranges[f'{letter}28'].value) == 'True':
            gui.bsym_nocom.set(True)
        else:
            gui.bsym_nocom.set(False)
        if str(sheet_ranges[f'{letter}29'].value) == 'True':
            gui.bsym_noint.set(True)
        else:
            gui.bsym_noint.set(False)
        if str(sheet_ranges[f'{letter}30'].value) == 'True':
            gui.bsym_nolaser.set(True)
        else:
            gui.bsym_nolaser.set(False)
        if str(sheet_ranges[f'{letter}31'].value) == 'True':
            gui.bsym_nopower.set(True)
        else:
            gui.bsym_nopower.set(False)
        if str(sheet_ranges[f'{letter}32'].value) == 'True':
            gui.bsym_symnff.set(True)
        else:
            gui.bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}33'].value) == 'True':
            gui.rp.set(True)
        else:
            gui.rp.set(False)
        if str(sheet_ranges[f'{letter}34'].value) == 'True':
            gui.nrp.set(True)
        else:
            gui.nrp.set(False)
        if str(sheet_ranges[f'{letter}35'].value) == 'True':
            gui.rt.set(True)
        else:
            gui.rt.set(False)
        if str(sheet_ranges[f'{letter}36'].value) == 'True':
            gui.kj.set(True)
        else:
            gui.kj.set(False)
        if str(sheet_ranges[f'{letter}37'].value) == 'True':
            gui.bhotswap.set(True)
        else:
            gui.bhotswap.set(False)
        if str(sheet_ranges[f'{letter}38'].value) == 'True':
            gui.bvibration.set(True)
        else:
            gui.bvibration.set(False)
        if str(sheet_ranges[f'{letter}39'].value) == 'True':
            gui.bnoise.set(True)
        else:
            gui.bnoise.set(False)
        if str(sheet_ranges[f'{letter}40'].value) == 'True':
            gui.bheat.set(True)
        else:
            gui.bheat.set(False)
        if str(sheet_ranges[f'{letter}41'].value) == 'True':
            gui.bnff.set(True)
        else:
            gui.bnff.set(False)
        if str(sheet_ranges[f'{letter}42'].value) == 'True':
            gui.bcable.set(True)
        else:
            gui.bcable.set(False)
        if str(sheet_ranges[f'{letter}43'].value) == 'True':
            gui.bsym_symnff.set(True)
        else:
            gui.bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}44'].value) == 'True':
            gui.bsym_cable.set(True)
        else:
            gui.bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}45'].value) == 'True':
            gui.bsurge.set(True)
        else:
            gui.bsurge.set(False)
        gui.c_appearance_text.insert(gui.tk.END, sheet_ranges[f'{letter}46'].value)
        gui.c_operation_text.insert(gui.tk.END, sheet_ranges[f'{letter}47'].value)
        gui.c_internal_text.insert(gui.tk.END, sheet_ranges[f'{letter}48'].value)
        gui.c_suspected_text.insert(gui.tk.END, sheet_ranges[f'{letter}49'].value)
        gui.c_conclusion_text.insert(gui.tk.END, sheet_ranges[f'{letter}50'].value)
        gui.bench_text.insert(gui.tk.END, sheet_ranges[f'{letter}51'].value)
        lb.close()
    elif gui.preset_text.get('1.0', 'end').split('\n')[0] == 'clear':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'A'
        gui.btn_preset_text.set('Preset 1')
        gui.preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}1'].value = 'Preset 1'
        for a in range(2, 51):
            sheet_ranges[f'{letter}{a}'].value = ''
        lb.save('presets.xlsx')
        lb.close()
    else:
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'A'

        sheet_ranges[f'{letter}1'].value = gui.preset_text.get('1.0', 'end').split('\n')[0]
        gui.btn_preset_text.set(sheet_ranges[f'{letter}1'].value)
        gui.preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}2'].value = gui.rpnum_text.get('1.0', 'end')
        sheet_ranges[f'{letter}3'].value = gui.item_text.get('1.0', 'end')
        # sheet_ranges[f'{letter}4'].value = serial_text.get('1.0', 'end')
        sheet_ranges[f'{letter}5'].value = gui.key_text.get('1.0', 'end')
        sheet_ranges[f'{letter}6'].value = gui.fb_text.get('1.0', 'end')
        sheet_ranges[f'{letter}7'].value = gui.comp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}8'].value = gui.structure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}9'].value = gui.parts_text.get('1.0', 'end')
        sheet_ranges[f'{letter}10'].value = gui.pic_text.get('1.0', 'end')
        sheet_ranges[f'{letter}11'].value = gui.his_rp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}12'].value = gui.his_comp_date_text.get('1.0', 'end')
        sheet_ranges[f'{letter}13'].value = gui.his_symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}14'].value = gui.his_result_text.get('1.0', 'end')
        sheet_ranges[f'{letter}15'].value = gui.appearance_text.get('1.0', 'end')
        sheet_ranges[f'{letter}16'].value = gui.symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}17'].value = gui.tests_text.get('1.0', 'end')
        sheet_ranges[f'{letter}18'].value = gui.analysis_text.get('1.0', 'end')
        sheet_ranges[f'{letter}19'].value = gui.failure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}20'].value = gui.bapp_impact.get()
        sheet_ranges[f'{letter}21'].value = gui.bapp_cable.get()
        sheet_ranges[f'{letter}22'].value = gui.bapp_filterglass.get()
        sheet_ranges[f'{letter}23'].value = gui.bapp_nff.get()
        sheet_ranges[f'{letter}24'].value = gui.bapp_oil.get()
        sheet_ranges[f'{letter}25'].value = gui.bsym_cable.get()
        sheet_ranges[f'{letter}26'].value = gui.bsym_inout.get()
        sheet_ranges[f'{letter}27'].value = gui.bsym_nffstruct.get()
        sheet_ranges[f'{letter}28'].value = gui.bsym_nocom.get()
        sheet_ranges[f'{letter}29'].value = gui.bsym_noint.get()
        sheet_ranges[f'{letter}30'].value = gui.bsym_nolaser.get()
        sheet_ranges[f'{letter}31'].value = gui.bsym_nopower.get()
        sheet_ranges[f'{letter}32'].value = gui.bsym_symnff.get()
        sheet_ranges[f'{letter}33'].value = gui.rp.get()
        sheet_ranges[f'{letter}34'].value = gui.nrp.get()
        sheet_ranges[f'{letter}35'].value = gui.rt.get()
        sheet_ranges[f'{letter}36'].value = gui.kj.get()
        sheet_ranges[f'{letter}37'].value = gui.bhotswap.get()
        sheet_ranges[f'{letter}38'].value = gui.bvibration.get()
        sheet_ranges[f'{letter}39'].value = gui.bnoise.get()
        sheet_ranges[f'{letter}40'].value = gui.bheat.get()
        sheet_ranges[f'{letter}41'].value = gui.bnff.get()
        sheet_ranges[f'{letter}42'].value = gui.bcable.get()
        sheet_ranges[f'{letter}43'].value = gui.bsym_symnff.get()
        sheet_ranges[f'{letter}44'].value = gui.bsym_cable.get()
        sheet_ranges[f'{letter}45'].value = gui.bsurge.get()
        sheet_ranges[f'{letter}46'].value = gui.c_appearance_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}47'].value = gui.c_operation_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}48'].value = gui.c_internal_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}49'].value = gui.c_suspected_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}50'].value = gui.c_conclusion_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}51'].value = gui.bench_text.get('1.0', 'end').split('\n')[0]
        lb.save('presets.xlsx')
        lb.close()
    return None


def save_exe_preset2():
    """Save all values into a preset 2."""
    if gui.preset_text.get('1.0', 'end').split('\n')[0] == '':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'B'
        gui.rpnum_text.insert(gui.tk.END, sheet_ranges[f'{letter}2'].value)
        gui.item_text.insert(gui.tk.END, sheet_ranges[f'{letter}3'].value)
        # serial_text.insert(tk.END, sheet_ranges[f'{letter}4'].value)
        gui.key_text.insert(gui.tk.END, sheet_ranges[f'{letter}5'].value)
        gui.fb_text.insert(gui.tk.END, sheet_ranges[f'{letter}6'].value)
        gui.comp_text.insert(gui.tk.END, sheet_ranges[f'{letter}7'].value)
        gui.structure_text.insert(gui.tk.END, sheet_ranges[f'{letter}8'].value)
        gui.parts_text.insert(gui.tk.END, sheet_ranges[f'{letter}9'].value)
        gui.pic_text.insert(gui.tk.END, sheet_ranges[f'{letter}10'].value)
        gui.his_rp_text.insert(gui.tk.END, sheet_ranges[f'{letter}11'].value)
        gui.his_comp_date_text.insert(gui.tk.END, sheet_ranges[f'{letter}12'].value)
        gui.his_symptom_text.insert(gui.tk.END, sheet_ranges[f'{letter}13'].value)
        gui.his_result_text.insert(gui.tk.END, sheet_ranges[f'{letter}14'].value)
        gui.appearance_text.insert(gui.tk.END, sheet_ranges[f'{letter}15'].value)
        gui.symptom_text.insert(gui.tk.END, sheet_ranges[f'{letter}16'].value)
        gui.tests_text.insert(gui.tk.END, sheet_ranges[f'{letter}17'].value)
        gui.analysis_text.insert(gui.tk.END, sheet_ranges[f'{letter}18'].value)
        gui.failure_text.insert(gui.tk.END, sheet_ranges[f'{letter}19'].value)
        if str(sheet_ranges[f'{letter}20'].value) == 'True':
            gui.bapp_impact.set(True)
        else:
            gui.bapp_impact.set(False)
        if str(sheet_ranges[f'{letter}21'].value) == 'True':
            gui.bapp_cable.set(True)
        else:
            gui.bapp_cable.set(False)
        if str(sheet_ranges[f'{letter}22'].value) == 'True':
            gui.bapp_filterglass.set(True)
        else:
            gui.bapp_filterglass.set(False)
        if str(sheet_ranges[f'{letter}23'].value) == 'True':
            gui.bapp_nff.set(True)
        else:
            gui.bapp_nff.set(False)
        if str(sheet_ranges[f'{letter}24'].value) == 'True':
            gui.bapp_oil.set(True)
        else:
            gui.bapp_oil.set(False)
        if str(sheet_ranges[f'{letter}25'].value) == 'True':
            gui.bsym_cable.set(True)
        else:
            gui.bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}26'].value) == 'True':
            gui.bsym_inout.set(True)
        else:
            gui.bsym_inout.set(False)
        if str(sheet_ranges[f'{letter}27'].value) == 'True':
            gui.bsym_nffstruct.set(True)
        else:
            gui.bsym_nffstruct.set(False)
        if str(sheet_ranges[f'{letter}28'].value) == 'True':
            gui.bsym_nocom.set(True)
        else:
            gui.bsym_nocom.set(False)
        if str(sheet_ranges[f'{letter}29'].value) == 'True':
            gui.bsym_noint.set(True)
        else:
            gui.bsym_noint.set(False)
        if str(sheet_ranges[f'{letter}30'].value) == 'True':
            gui.bsym_nolaser.set(True)
        else:
            gui.bsym_nolaser.set(False)
        if str(sheet_ranges[f'{letter}31'].value) == 'True':
            gui.bsym_nopower.set(True)
        else:
            gui.bsym_nopower.set(False)
        if str(sheet_ranges[f'{letter}32'].value) == 'True':
            gui.bsym_symnff.set(True)
        else:
            gui.bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}33'].value) == 'True':
            gui.rp.set(True)
        else:
            gui.rp.set(False)
        if str(sheet_ranges[f'{letter}34'].value) == 'True':
            gui.nrp.set(True)
        else:
            gui.nrp.set(False)
        if str(sheet_ranges[f'{letter}35'].value) == 'True':
            gui.rt.set(True)
        else:
            gui.rt.set(False)
        if str(sheet_ranges[f'{letter}36'].value) == 'True':
            gui.kj.set(True)
        else:
            gui.kj.set(False)
        if str(sheet_ranges[f'{letter}37'].value) == 'True':
            gui.bhotswap.set(True)
        else:
            gui.bhotswap.set(False)
        if str(sheet_ranges[f'{letter}38'].value) == 'True':
            gui.bvibration.set(True)
        else:
            gui.bvibration.set(False)
        if str(sheet_ranges[f'{letter}39'].value) == 'True':
            gui.bnoise.set(True)
        else:
            gui.bnoise.set(False)
        if str(sheet_ranges[f'{letter}40'].value) == 'True':
            gui.bheat.set(True)
        else:
            gui.bheat.set(False)
        if str(sheet_ranges[f'{letter}41'].value) == 'True':
            gui.bnff.set(True)
        else:
            gui.bnff.set(False)
        if str(sheet_ranges[f'{letter}42'].value) == 'True':
            gui.bcable.set(True)
        else:
            gui.bcable.set(False)
        if str(sheet_ranges[f'{letter}43'].value) == 'True':
            gui.bsym_symnff.set(True)
        else:
            gui.bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}44'].value) == 'True':
            gui.bsym_cable.set(True)
        else:
            gui.bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}45'].value) == 'True':
            gui.bsurge.set(True)
        else:
            gui.bsurge.set(False)
        gui.c_appearance_text.insert(gui.tk.END, sheet_ranges[f'{letter}46'].value)
        gui.c_operation_text.insert(gui.tk.END, sheet_ranges[f'{letter}47'].value)
        gui.c_internal_text.insert(gui.tk.END, sheet_ranges[f'{letter}48'].value)
        gui.c_suspected_text.insert(gui.tk.END, sheet_ranges[f'{letter}49'].value)
        gui.c_conclusion_text.insert(gui.tk.END, sheet_ranges[f'{letter}50'].value)
        gui.bench_text.insert(gui.tk.END, sheet_ranges[f'{letter}51'].value)
        lb.close()
    elif gui.preset_text.get('1.0', 'end').split('\n')[0] == 'clear':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'B'
        gui.btn_preset2_text.set('Preset 2')
        gui.preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}1'].value = 'Preset 2'
        for a in range(2, 51):
            sheet_ranges[f'{letter}{a}'].value = ''
        lb.save('presets.xlsx')
        lb.close()
    else:
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'B'

        sheet_ranges[f'{letter}1'].value = gui.preset_text.get('1.0', 'end').split('\n')[0]
        gui.btn_preset2_text.set(sheet_ranges[f'{letter}1'].value)
        gui.preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}2'].value = gui.rpnum_text.get('1.0', 'end')
        sheet_ranges[f'{letter}3'].value = gui.item_text.get('1.0', 'end')
        # sheet_ranges[f'{letter}4'].value = serial_text.get('1.0', 'end')
        sheet_ranges[f'{letter}5'].value = gui.key_text.get('1.0', 'end')
        sheet_ranges[f'{letter}6'].value = gui.fb_text.get('1.0', 'end')
        sheet_ranges[f'{letter}7'].value = gui.comp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}8'].value = gui.structure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}9'].value = gui.parts_text.get('1.0', 'end')
        sheet_ranges[f'{letter}10'].value = gui.pic_text.get('1.0', 'end')
        sheet_ranges[f'{letter}11'].value = gui.his_rp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}12'].value = gui.his_comp_date_text.get('1.0', 'end')
        sheet_ranges[f'{letter}13'].value = gui.his_symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}14'].value = gui.his_result_text.get('1.0', 'end')
        sheet_ranges[f'{letter}15'].value = gui.appearance_text.get('1.0', 'end')
        sheet_ranges[f'{letter}16'].value = gui.symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}17'].value = gui.tests_text.get('1.0', 'end')
        sheet_ranges[f'{letter}18'].value = gui.analysis_text.get('1.0', 'end')
        sheet_ranges[f'{letter}19'].value = gui.failure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}20'].value = gui.bapp_impact.get()
        sheet_ranges[f'{letter}21'].value = gui.bapp_cable.get()
        sheet_ranges[f'{letter}22'].value = gui.bapp_filterglass.get()
        sheet_ranges[f'{letter}23'].value = gui.bapp_nff.get()
        sheet_ranges[f'{letter}24'].value = gui.bapp_oil.get()
        sheet_ranges[f'{letter}25'].value = gui.bsym_cable.get()
        sheet_ranges[f'{letter}26'].value = gui.bsym_inout.get()
        sheet_ranges[f'{letter}27'].value = gui.bsym_nffstruct.get()
        sheet_ranges[f'{letter}28'].value = gui.bsym_nocom.get()
        sheet_ranges[f'{letter}29'].value = gui.bsym_noint.get()
        sheet_ranges[f'{letter}30'].value = gui.bsym_nolaser.get()
        sheet_ranges[f'{letter}31'].value = gui.bsym_nopower.get()
        sheet_ranges[f'{letter}32'].value = gui.bsym_symnff.get()
        sheet_ranges[f'{letter}33'].value = gui.rp.get()
        sheet_ranges[f'{letter}34'].value = gui.nrp.get()
        sheet_ranges[f'{letter}35'].value = gui.rt.get()
        sheet_ranges[f'{letter}36'].value = gui.kj.get()
        sheet_ranges[f'{letter}37'].value = gui.bhotswap.get()
        sheet_ranges[f'{letter}38'].value = gui.bvibration.get()
        sheet_ranges[f'{letter}39'].value = gui.bnoise.get()
        sheet_ranges[f'{letter}40'].value = gui.bheat.get()
        sheet_ranges[f'{letter}41'].value = gui.bnff.get()
        sheet_ranges[f'{letter}42'].value = gui.bcable.get()
        sheet_ranges[f'{letter}43'].value = gui.bsym_symnff.get()
        sheet_ranges[f'{letter}44'].value = gui.bsym_cable.get()
        sheet_ranges[f'{letter}45'].value = gui.bsurge.get()
        sheet_ranges[f'{letter}46'].value = gui.c_appearance_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}47'].value = gui.c_operation_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}48'].value = gui.c_internal_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}49'].value = gui.c_suspected_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}50'].value = gui.c_conclusion_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}51'].value = gui.bench_text.get('1.0', 'end').split('\n')[0]
        lb.save('presets.xlsx')
        lb.close()
    return None


def save_exe_preset3():
    """Save all values into a preset 3."""
    if gui.preset_text.get('1.0', 'end').split('\n')[0] == '':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'C'
        gui.rpnum_text.insert(gui.tk.END, sheet_ranges[f'{letter}2'].value)
        gui.item_text.insert(gui.tk.END, sheet_ranges[f'{letter}3'].value)
        # serial_text.insert(tk.END, sheet_ranges[f'{letter}4'].value)
        gui.key_text.insert(gui.tk.END, sheet_ranges[f'{letter}5'].value)
        gui.fb_text.insert(gui.tk.END, sheet_ranges[f'{letter}6'].value)
        gui.comp_text.insert(gui.tk.END, sheet_ranges[f'{letter}7'].value)
        gui.structure_text.insert(gui.tk.END, sheet_ranges[f'{letter}8'].value)
        gui.parts_text.insert(gui.tk.END, sheet_ranges[f'{letter}9'].value)
        gui.pic_text.insert(gui.tk.END, sheet_ranges[f'{letter}10'].value)
        gui.his_rp_text.insert(gui.tk.END, sheet_ranges[f'{letter}11'].value)
        gui.his_comp_date_text.insert(gui.tk.END, sheet_ranges[f'{letter}12'].value)
        gui.his_symptom_text.insert(gui.tk.END, sheet_ranges[f'{letter}13'].value)
        gui.his_result_text.insert(gui.tk.END, sheet_ranges[f'{letter}14'].value)
        gui.appearance_text.insert(gui.tk.END, sheet_ranges[f'{letter}15'].value)
        gui.symptom_text.insert(gui.tk.END, sheet_ranges[f'{letter}16'].value)
        gui.tests_text.insert(gui.tk.END, sheet_ranges[f'{letter}17'].value)
        gui.analysis_text.insert(gui.tk.END, sheet_ranges[f'{letter}18'].value)
        gui.failure_text.insert(gui.tk.END, sheet_ranges[f'{letter}19'].value)
        if str(sheet_ranges[f'{letter}20'].value) == 'True':
            gui.bapp_impact.set(True)
        else:
            gui.bapp_impact.set(False)
        if str(sheet_ranges[f'{letter}21'].value) == 'True':
            gui.bapp_cable.set(True)
        else:
            gui.bapp_cable.set(False)
        if str(sheet_ranges[f'{letter}22'].value) == 'True':
            gui.bapp_filterglass.set(True)
        else:
            gui.bapp_filterglass.set(False)
        if str(sheet_ranges[f'{letter}23'].value) == 'True':
            gui.bapp_nff.set(True)
        else:
            gui.bapp_nff.set(False)
        if str(sheet_ranges[f'{letter}24'].value) == 'True':
            gui.bapp_oil.set(True)
        else:
            gui.bapp_oil.set(False)
        if str(sheet_ranges[f'{letter}25'].value) == 'True':
            gui.bsym_cable.set(True)
        else:
            gui.bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}26'].value) == 'True':
            gui.bsym_inout.set(True)
        else:
            gui.bsym_inout.set(False)
        if str(sheet_ranges[f'{letter}27'].value) == 'True':
            gui.bsym_nffstruct.set(True)
        else:
            gui.bsym_nffstruct.set(False)
        if str(sheet_ranges[f'{letter}28'].value) == 'True':
            gui.bsym_nocom.set(True)
        else:
            gui.bsym_nocom.set(False)
        if str(sheet_ranges[f'{letter}29'].value) == 'True':
            gui.bsym_noint.set(True)
        else:
            gui.bsym_noint.set(False)
        if str(sheet_ranges[f'{letter}30'].value) == 'True':
            gui.bsym_nolaser.set(True)
        else:
            gui.bsym_nolaser.set(False)
        if str(sheet_ranges[f'{letter}31'].value) == 'True':
            gui.bsym_nopower.set(True)
        else:
            gui.bsym_nopower.set(False)
        if str(sheet_ranges[f'{letter}32'].value) == 'True':
            gui.bsym_symnff.set(True)
        else:
            gui.bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}33'].value) == 'True':
            gui.rp.set(True)
        else:
            gui.rp.set(False)
        if str(sheet_ranges[f'{letter}34'].value) == 'True':
            gui.nrp.set(True)
        else:
            gui.nrp.set(False)
        if str(sheet_ranges[f'{letter}35'].value) == 'True':
            gui.rt.set(True)
        else:
            gui.rt.set(False)
        if str(sheet_ranges[f'{letter}36'].value) == 'True':
            gui.kj.set(True)
        else:
            gui.kj.set(False)
        if str(sheet_ranges[f'{letter}37'].value) == 'True':
            gui.bhotswap.set(True)
        else:
            gui.bhotswap.set(False)
        if str(sheet_ranges[f'{letter}38'].value) == 'True':
            gui.bvibration.set(True)
        else:
            gui.bvibration.set(False)
        if str(sheet_ranges[f'{letter}39'].value) == 'True':
            gui.bnoise.set(True)
        else:
            gui.bnoise.set(False)
        if str(sheet_ranges[f'{letter}40'].value) == 'True':
            gui.bheat.set(True)
        else:
            gui.bheat.set(False)
        if str(sheet_ranges[f'{letter}41'].value) == 'True':
            gui.bnff.set(True)
        else:
            gui.bnff.set(False)
        if str(sheet_ranges[f'{letter}42'].value) == 'True':
            gui.bcable.set(True)
        else:
            gui.bcable.set(False)
        if str(sheet_ranges[f'{letter}43'].value) == 'True':
            gui.bsym_symnff.set(True)
        else:
            gui.bsym_symnff.set(False)
        if str(sheet_ranges[f'{letter}44'].value) == 'True':
            gui.bsym_cable.set(True)
        else:
            gui.bsym_cable.set(False)
        if str(sheet_ranges[f'{letter}45'].value) == 'True':
            gui.bsurge.set(True)
        else:
            gui.bsurge.set(False)
        gui.c_appearance_text.insert(gui.tk.END, sheet_ranges[f'{letter}46'].value)
        gui.c_operation_text.insert(gui.tk.END, sheet_ranges[f'{letter}47'].value)
        gui.c_internal_text.insert(gui.tk.END, sheet_ranges[f'{letter}48'].value)
        gui.c_suspected_text.insert(gui.tk.END, sheet_ranges[f'{letter}49'].value)
        gui.c_conclusion_text.insert(gui.tk.END, sheet_ranges[f'{letter}50'].value)
        gui.bench_text.insert(gui.tk.END, sheet_ranges[f'{letter}51'].value)
        lb.close()
    elif gui.preset_text.get('1.0', 'end').split('\n')[0] == 'clear':
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'C'
        gui.btn_preset3_text.set('Preset 3')
        gui.preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}1'].value = 'Preset 3'
        for a in range(2, 51):
            sheet_ranges[f'{letter}{a}'].value = ''
        lb.save('presets.xlsx')
        lb.close()
    else:
        lb = load_workbook(filename='presets.xlsx')
        sheet_ranges = lb['Sheet']
        letter = 'C'

        sheet_ranges[f'{letter}1'].value = gui.preset_text.get('1.0', 'end').split('\n')[0]
        gui.btn_preset3_text.set(sheet_ranges[f'{letter}1'].value)
        gui.preset_text.delete('1.0', 'end')

        sheet_ranges[f'{letter}2'].value = gui.rpnum_text.get('1.0', 'end')
        sheet_ranges[f'{letter}3'].value = gui.item_text.get('1.0', 'end')
        # sheet_ranges[f'{letter}4'].value = serial_text.get('1.0', 'end')
        sheet_ranges[f'{letter}5'].value = gui.key_text.get('1.0', 'end')
        sheet_ranges[f'{letter}6'].value = gui.fb_text.get('1.0', 'end')
        sheet_ranges[f'{letter}7'].value = gui.comp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}8'].value = gui.structure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}9'].value = gui.parts_text.get('1.0', 'end')
        sheet_ranges[f'{letter}10'].value = gui.pic_text.get('1.0', 'end')
        sheet_ranges[f'{letter}11'].value = gui.his_rp_text.get('1.0', 'end')
        sheet_ranges[f'{letter}12'].value = gui.his_comp_date_text.get('1.0', 'end')
        sheet_ranges[f'{letter}13'].value = gui.his_symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}14'].value = gui.his_result_text.get('1.0', 'end')
        sheet_ranges[f'{letter}15'].value = gui.appearance_text.get('1.0', 'end')
        sheet_ranges[f'{letter}16'].value = gui.symptom_text.get('1.0', 'end')
        sheet_ranges[f'{letter}17'].value = gui.tests_text.get('1.0', 'end')
        sheet_ranges[f'{letter}18'].value = gui.analysis_text.get('1.0', 'end')
        sheet_ranges[f'{letter}19'].value = gui.failure_text.get('1.0', 'end')
        sheet_ranges[f'{letter}20'].value = gui.bapp_impact.get()
        sheet_ranges[f'{letter}21'].value = gui.bapp_cable.get()
        sheet_ranges[f'{letter}22'].value = gui.bapp_filterglass.get()
        sheet_ranges[f'{letter}23'].value = gui.bapp_nff.get()
        sheet_ranges[f'{letter}24'].value = gui.bapp_oil.get()
        sheet_ranges[f'{letter}25'].value = gui.bsym_cable.get()
        sheet_ranges[f'{letter}26'].value = gui.bsym_inout.get()
        sheet_ranges[f'{letter}27'].value = gui.bsym_nffstruct.get()
        sheet_ranges[f'{letter}28'].value = gui.bsym_nocom.get()
        sheet_ranges[f'{letter}29'].value = gui.bsym_noint.get()
        sheet_ranges[f'{letter}30'].value = gui.bsym_nolaser.get()
        sheet_ranges[f'{letter}31'].value = gui.bsym_nopower.get()
        sheet_ranges[f'{letter}32'].value = gui.bsym_symnff.get()
        sheet_ranges[f'{letter}33'].value = gui.rp.get()
        sheet_ranges[f'{letter}34'].value = gui.nrp.get()
        sheet_ranges[f'{letter}35'].value = gui.rt.get()
        sheet_ranges[f'{letter}36'].value = gui.kj.get()
        sheet_ranges[f'{letter}37'].value = gui.bhotswap.get()
        sheet_ranges[f'{letter}38'].value = gui.bvibration.get()
        sheet_ranges[f'{letter}39'].value = gui.bnoise.get()
        sheet_ranges[f'{letter}40'].value = gui.bheat.get()
        sheet_ranges[f'{letter}41'].value = gui.bnff.get()
        sheet_ranges[f'{letter}42'].value = gui.bcable.get()
        sheet_ranges[f'{letter}43'].value = gui.bsym_symnff.get()
        sheet_ranges[f'{letter}44'].value = gui.bsym_cable.get()
        sheet_ranges[f'{letter}45'].value = gui.bsurge.get()
        sheet_ranges[f'{letter}46'].value = gui.c_appearance_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}47'].value = gui.c_operation_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}48'].value = gui.c_internal_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}49'].value = gui.c_suspected_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}50'].value = gui.c_conclusion_text.get('1.0', 'end').split('\n')[0]
        sheet_ranges[f'{letter}51'].value = gui.bench_text.get('1.0', 'end').split('\n')[0]
        lb.save('presets.xlsx')
        lb.close()
    return None
